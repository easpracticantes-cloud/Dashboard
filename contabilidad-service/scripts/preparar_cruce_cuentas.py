"""
Convierte el Excel de trabajo «CRUCE DE CUENTAS» a un archivo
listo para importar en Sistema Contable IA (Autobits).

Uso:
  python scripts/preparar_cruce_cuentas.py "C:\\ruta\\CRUCE DE CUENTAS 2026.xlsx"
  python scripts/preparar_cruce_cuentas.py   # usa la ruta por defecto en Descargas
"""

from __future__ import annotations

import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

# Permitir importar desde src/
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from domain.autobits.fields import AUTOBITS_EXPORT_COLUMNS  # noqa: E402

DEFAULT_INPUT = Path(r"C:\Users\07sam\Downloads\CRUCE DE CUENTAS 2026.xlsx")
DEFAULT_OUTPUT = ROOT / "dataset" / "demo" / "cruce_cuentas_LISTO_PARA_IMPORTAR.xlsx"

HEADER_ALIASES = {
    "fecha": ["fecha de ejecución", "fecha de ejecucion", "fecha"],
    "compra": ["orden de compra", "codigo orden de compra", "código orden de compra"],
    "reserva": ["ref.", "ref", "codigo reserva", "código reserva", "referencia"],
    "valor": ["valor", "total", "precio eas", "precio terceros"],
    "factura": ["factura/cdc", "factura", "cdc"],
    "pago": ["fecha de pago"],
    "nit": ["nit/cc proveedor", "nit"],
    "proveedor": ["nombre proveedor"],
    "concepto": ["nombre concepto", "description servicio", "descripcion"],
    "estado": ["estado de la compra"],
}


def _norm(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def _to_date(value) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text[:19], fmt).date().isoformat()
        except ValueError:
            continue
    # fechas rotas tipo 12/0772026
    m = re.match(r"(\d{1,2})[/-](\d{1,2})[/-]?(\d{4})", text)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if mo > 12:
            mo = int(str(mo)[:2]) if len(str(mo)) > 2 else mo
        try:
            return date(y, mo, d).isoformat()
        except ValueError:
            return None
    return None


def _to_float(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("$", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _to_str(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _match_field(header: str) -> str | None:
    h = _norm(header)
    if not h:
        return None
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if h == alias or h.startswith(alias) or alias in h:
                return field
    return None


def _is_header_row(values: list) -> bool:
    fields = {_match_field(v) for v in values if v is not None}
    return "compra" in fields and ("valor" in fields or "fecha" in fields)


def _extract_blocks(header_values: list) -> list[dict]:
    """Detecta bloques de columnas a partir de una fila de encabezados."""
    indexed = []
    for idx, cell in enumerate(header_values):
        field = _match_field(cell) if cell is not None else None
        if field:
            indexed.append((idx, field))
    if not indexed:
        return []

    blocks: list[dict] = []
    current: dict[str, int] = {}
    last_idx = None
    for idx, field in indexed:
        # nuevo bloque si hay salto grande o se repite un campo clave
        if last_idx is not None and (idx - last_idx > 2 or field in current):
            if "compra" in current or "valor" in current:
                blocks.append(current)
            current = {}
        current[field] = idx
        last_idx = idx
    if current and ("compra" in current or "valor" in current):
        blocks.append(current)
    return blocks


JUNK_PROVIDERS = {
    "impresa",
    "unidades disponibles año 2026",
    "unidades disponibles ano 2026",
    "total pagados",
    "entradas a favor",
    "saldo anterior al noviembre 2025",
    "si",
    "no",
}


def _looks_like_invoice(text: str) -> bool:
    t = _norm(text)
    return bool(
        re.search(r"\b(fv|fe|fpos|cdc|factura|pos)\b", t)
        or re.match(r"^(fv|fe|fpos|cdc)\b", t)
        or t in {"impresa"}
    )


def _looks_like_provider_name(text: str) -> bool:
    if not text or len(text.strip()) < 4:
        return False
    t = text.strip()
    if _match_field(t):
        return False
    if _to_date(t) or _to_float(t) is not None:
        return False
    if _looks_like_invoice(t):
        return False
    if _norm(t) in JUNK_PROVIDERS:
        return False
    if "unidades disponibles" in _norm(t):
        return False
    upper = t.upper()
    if upper.startswith(("COM", "EAS", "COT")):
        return False
    letters = sum(ch.isalpha() for ch in t)
    return letters >= 4

def _provider_above(rows: list[list], header_row_idx: int, block: dict) -> str | None:
    cols = list(block.values())
    c0, c1 = min(cols), max(cols)
    for r in range(header_row_idx - 1, max(-1, header_row_idx - 8), -1):
        if r < 0:
            break
        for c in range(max(0, c0 - 1), min(len(rows[r]), c1 + 2)):
            val = rows[r][c]
            if val is None:
                continue
            text = str(val).strip()
            if _looks_like_provider_name(text):
                return text
    return None


def _row_from_block(row: list, block: dict, proveedor: str | None, hoja: str) -> dict | None:
    def get(field: str):
        idx = block.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    compra = _to_str(get("compra"))
    valor = _to_float(get("valor"))
    reserva = _to_str(get("reserva"))
    fecha = _to_date(get("fecha"))
    factura = _to_str(get("factura"))
    concepto = _to_str(get("concepto"))
    nit = _to_str(get("nit"))
    prov_cell = _to_str(get("proveedor"))
    prov = prov_cell if (prov_cell and _looks_like_provider_name(prov_cell)) else proveedor

    if compra and _norm(compra) in {"orden de compra", "codigo orden de compra", "código orden de compra"}:
        return None
    if reserva and _norm(reserva) in {"ref.", "ref", "codigo reserva"}:
        return None

    # Fila útil: necesita compra o (valor + reserva/factura)
    if not compra and not (valor is not None and (reserva or factura)):
        return None
    if not compra and not valor:
        return None

    if prov and _looks_like_invoice(prov):
        prov = None

    return {
        "NIT/CC Proveedor (Orden de Compra)": nit,
        "Nombre Proveedor (Orden de Compra)": prov,
        "Codigo Orden de compra": compra,
        "Codigo Reserva": reserva,
        "Fecha de ejecución (Reserva)": fecha,
        "estado de la compra": _to_str(get("estado")) or "",
        "Nombre concepto": concepto or "",
        "Moneda": "COP",
        "Total": valor,
        "SI": "",
        "NO": "",
        "OBSERVACIONES": f"Hoja:{hoja}" + (f" | Factura:{factura}" if factura else ""),
    }


def parse_sheet_matrix(ws) -> list[list]:
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def parse_cruce_sheet(name: str, rows: list[list]) -> list[dict]:
    """Parsea hojas tipo cruce (proveedores en bloques laterales)."""
    out: list[dict] = []
    active_blocks: list[tuple[dict, str | None]] = []

    for r_idx, row in enumerate(rows):
        if _is_header_row(row):
            blocks = _extract_blocks(row)
            active_blocks = [(b, _provider_above(rows, r_idx, b)) for b in blocks]
            # si en la misma fila hay nombres de proveedor sueltos, actualizar
            continue

        # nombre de proveedor intercalado (ej. nueva sección mid-sheet)
        for c, cell in enumerate(row):
            if cell is None:
                continue
            text = str(cell).strip()
            if not _looks_like_provider_name(text):
                continue
            for i, (block, _prov) in enumerate(active_blocks):
                cols = list(block.values())
                if min(cols) - 1 <= c <= max(cols) + 1:
                    active_blocks[i] = (block, text)

        for block, proveedor in active_blocks:
            item = _row_from_block(row, block, proveedor, name)
            if item:
                out.append(item)
    return out


def parse_autobits_like_sheet(name: str, rows: list[list]) -> list[dict]:
    """Parsea hojas con encabezados tipo export Autobits."""
    header_idx = None
    mapping: dict[str, int] = {}
    for i, row in enumerate(rows):
        fields = {}
        for idx, cell in enumerate(row):
            f = _match_field(cell) if cell is not None else None
            if f and f not in fields:
                fields[f] = idx
        if "compra" in fields and ("nit" in fields or "proveedor" in fields or "valor" in fields):
            header_idx = i
            mapping = fields
            break
    if header_idx is None:
        return []

    out = []
    for row in rows[header_idx + 1 :]:
        item = _row_from_block(row, mapping, None, name)
        if item:
            # VENTAS_DUSTER: valor preferir PRECIO EAS si existe (ya mapeado a valor)
            out.append(item)
    return out


def convert_workbook(input_path: Path) -> list[dict]:
    wb = load_workbook(input_path, read_only=True, data_only=True)
    all_rows: list[dict] = []
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            matrix = parse_sheet_matrix(ws)
            # Preferir parser Autobits-like si hay NIT/Nombre Proveedor
            flat = " ".join(_norm(c) for row in matrix[:15] for c in row if c is not None)
            if "precompra" in _norm(name):
                parsed = parse_cruce_sheet(name, matrix)
                for item in parsed:
                    if not item.get("Nombre Proveedor (Orden de Compra)"):
                        item["Nombre Proveedor (Orden de Compra)"] = "PRECOMPRA LUGER"
            elif "nit/cc proveedor" in flat or "nombre proveedor (orden de compra)" in flat:
                parsed = parse_autobits_like_sheet(name, matrix)
            else:
                parsed = parse_cruce_sheet(name, matrix)
            # limpiar proveedores basura
            clean = []
            for item in parsed:
                prov = item.get("Nombre Proveedor (Orden de Compra)")
                if prov and _norm(prov) in JUNK_PROVIDERS:
                    item["Nombre Proveedor (Orden de Compra)"] = None
                if prov and "unidades disponibles" in _norm(str(prov)):
                    item["Nombre Proveedor (Orden de Compra)"] = "PRECOMPRA LUGER"
                clean.append(item)
            parsed = clean
            print(f"  · Hoja «{name}»: {len(parsed)} fila(s)")
            all_rows.extend(parsed)
    finally:
        wb.close()
    return all_rows


def dedupe(rows: list[dict]) -> list[dict]:
    seen = set()
    out = []
    for row in rows:
        key = (
            row.get("NIT/CC Proveedor (Orden de Compra)"),
            row.get("Nombre Proveedor (Orden de Compra)"),
            row.get("Codigo Orden de compra"),
            row.get("Codigo Reserva"),
            row.get("Total"),
            row.get("Fecha de ejecución (Reserva)"),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out


def write_output(rows: list[dict], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Autobits"
    ws.append(list(AUTOBITS_EXPORT_COLUMNS))
    for row in rows:
        ws.append([row.get(col) for col in AUTOBITS_EXPORT_COLUMNS])
    wb.save(output_path)
    return output_path


def main(argv: list[str] | None = None) -> int:
    args = argv if argv is not None else sys.argv[1:]
    input_path = Path(args[0]) if args else DEFAULT_INPUT
    output_path = Path(args[1]) if len(args) > 1 else DEFAULT_OUTPUT

    if not input_path.exists():
        print(f"No se encontró el archivo: {input_path}")
        return 1

    print(f"Leyendo: {input_path}")
    rows = convert_workbook(input_path)
    rows = dedupe(rows)
    print(f"Total filas útiles (sin duplicados): {len(rows)}")
    if not rows:
        print("No se extrajeron filas. Revise el formato del Excel.")
        return 2

    out = write_output(rows, output_path)
    # Copia fácil en Descargas
    copy = Path(r"C:\Users\07sam\Downloads") / "cruce_cuentas_LISTO_PARA_IMPORTAR.xlsx"
    write_output(rows, copy)
    print()
    print("LISTO PARA IMPORTAR:")
    print(f"  1) {out}")
    print(f"  2) {copy}")
    print()
    print("En la app: Autobits > Adjuntar Excel > elegir ese archivo.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
