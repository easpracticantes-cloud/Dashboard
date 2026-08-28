"""Adaptador Excel para importación de reportes Autobits."""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from domain.autobits.fields import (
    AUTOBITS_FIELDS,
    ParsedAutobitsRow,
    suggest_mapping,
)
from openpyxl import load_workbook


class AutobitsImportError(Exception):
    def __init__(self, message: str, code: str = "IMPORT_ERROR"):
        super().__init__(message)
        self.message = message
        self.code = code


@dataclass
class ExcelPreviewResult:
    columns: list[str]
    sample_rows: list[dict]
    suggested_mapping: dict[str, str | None]
    total_rows: int
    sheet_name: str


@dataclass
class ExcelParseResult:
    rows: list[ParsedAutobitsRow] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    skipped_empty: int = 0


def _to_float(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        texto = str(value).strip().replace("$", "").replace(" ", "")
        if "," in texto and "." in texto:
            texto = texto.replace(".", "").replace(",", ".")
        elif "," in texto:
            texto = texto.replace(",", ".")
        return float(texto)
    except ValueError:
        return None


def _to_date_str(value) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    texto = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(texto, fmt).date().isoformat()
        except ValueError:
            continue
    return texto or None


def _to_str(value) -> str | None:
    if value is None:
        return None
    texto = str(value).strip()
    return texto or None


def _read_sheet(path: Path):
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise AutobitsImportError(
            "Formato no soportado. Use un archivo Excel (.xlsx).",
            "INVALID_FORMAT",
        )
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise AutobitsImportError(f"No se pudo leer el Excel: {exc}", "READ_ERROR") from exc
    sheet = wb.active
    if sheet is None:
        wb.close()
        raise AutobitsImportError("El archivo Excel no tiene hojas.", "EMPTY_WORKBOOK")
    return wb, sheet


def _iter_data_rows(sheet) -> tuple[list[str], list[tuple[int, dict]]]:
    """Lee filas; busca la fila de encabezados real (Autobits a veces pone título/copyright arriba)."""
    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        return [], []

    header_hints = (
        "proveedor",
        "nit",
        "orden",
        "compra",
        "reserva",
        "total",
        "observacion",
        "concepto",
        "fecha",
    )

    def score_header(row) -> int:
        score = 0
        for cell in row:
            if cell is None:
                continue
            t = str(cell).strip().lower()
            if not t or t.startswith("columna_"):
                continue
            if any(h in t for h in header_hints):
                score += 2
            elif not str(cell).replace(".", "").isdigit() and len(t) > 2:
                score += 1
        return score

    best_idx = 0
    best_score = -1
    for i, row in enumerate(all_rows[:20]):
        sc = score_header(row or ())
        if sc > best_score:
            best_score = sc
            best_idx = i

    # Si la "mejor" fila no parece encabezado, usar la primera
    if best_score < 4:
        best_idx = 0

    header_row = all_rows[best_idx]
    columns: list[str] = []
    for idx, cell in enumerate(header_row):
        label = _to_str(cell) or f"Columna_{idx + 1}"
        # Evitar copyright/título como nombre de columna
        low = label.lower()
        if "all rights reserved" in low or "autobits systems" in low:
            label = f"Columna_{idx + 1}"
        columns.append(label)

    # Desambiguar nombres duplicados
    seen: dict[str, int] = {}
    unique_cols: list[str] = []
    for col in columns:
        if col not in seen:
            seen[col] = 0
            unique_cols.append(col)
        else:
            seen[col] += 1
            unique_cols.append(f"{col}_{seen[col]}")
    columns = unique_cols

    data_rows: list[tuple[int, dict]] = []
    for offset, row in enumerate(all_rows[best_idx + 1 :], start=best_idx + 2):
        row_dict = {}
        has_value = False
        for col_name, cell in zip(columns, row or ()):
            if cell is not None and str(cell).strip() != "":
                has_value = True
            row_dict[col_name] = cell
        if has_value:
            data_rows.append((offset, row_dict))
    return columns, data_rows


class ExcelAutobitsAdapter:
    """Importador v1 de reportes semanales Autobits desde Excel."""

    def preview(self, path: Path, sample_limit: int = 5) -> ExcelPreviewResult:
        wb, sheet = _read_sheet(path)
        try:
            columns, data_rows = _iter_data_rows(sheet)
            if not columns:
                raise AutobitsImportError("El Excel está vacío o sin encabezados.", "EMPTY_SHEET")

            sample = []
            for _, row_dict in data_rows[:sample_limit]:
                sample.append({k: _serialize_cell(v) for k, v in row_dict.items()})

            return ExcelPreviewResult(
                columns=columns,
                sample_rows=sample,
                suggested_mapping=suggest_mapping(columns),
                total_rows=len(data_rows),
                sheet_name=sheet.title or "Sheet1",
            )
        finally:
            wb.close()

    def parse(
        self,
        path: Path,
        mapping: dict[str, str | None],
        *,
        validate: bool = True,
    ) -> ExcelParseResult:
        wb, sheet = _read_sheet(path)
        try:
            columns, data_rows = _iter_data_rows(sheet)
            if not columns:
                raise AutobitsImportError("El Excel está vacío o sin encabezados.", "EMPTY_SHEET")

            active_mapping = {k: v for k, v in mapping.items() if v and v in columns}
            if validate and not active_mapping.get("valor") and not active_mapping.get("proveedor"):
                raise AutobitsImportError(
                    "Debe mapear al menos Proveedor o Valor.",
                    "INVALID_MAPPING",
                )

            result = ExcelParseResult()
            for excel_row, row_dict in data_rows:
                parsed = self._parse_row(excel_row, row_dict, active_mapping)
                if parsed.is_empty():
                    result.skipped_empty += 1
                    continue
                if parsed.errors:
                    result.errors.extend(parsed.errors)
                result.rows.append(parsed)
            return result
        finally:
            wb.close()

    def _parse_row(
        self,
        row_number: int,
        row_dict: dict,
        mapping: dict[str, str | None],
    ) -> ParsedAutobitsRow:
        errors: list[str] = []

        def get(field: str):
            col = mapping.get(field)
            if not col:
                return None
            return row_dict.get(col)

        valor = _to_float(get("valor"))
        if mapping.get("valor") and get("valor") not in (None, "") and valor is None:
            errors.append(f"Fila {row_number}: valor inválido")

        raw = {k: _serialize_cell(v) for k, v in row_dict.items()}
        observaciones = _to_str(get("observaciones"))
        estado_compra = _to_str(get("estado_compra"))
        if not observaciones:
            from domain.autobits.observaciones import extract_observaciones_from_raw

            observaciones = extract_observaciones_from_raw(raw)
        if not estado_compra:
            from domain.autobits.observaciones import extract_estado_compra_from_raw

            estado_compra = extract_estado_compra_from_raw(raw)

        parsed = ParsedAutobitsRow(
            row_number=row_number,
            proveedor=_to_str(get("proveedor")),
            nit=_to_str(get("nit")),
            numero_compra=_to_str(get("numero_compra")),
            numero_reserva=_to_str(get("numero_reserva")),
            numero_documento=_to_str(get("numero_documento")),
            valor=valor,
            fecha=_to_date_str(get("fecha")),
            concepto=_to_str(get("concepto")),
            observaciones=observaciones,
            estado_compra=estado_compra,
            raw=raw,
            errors=errors or None,
        )
        return parsed

    def export_rows_csv(self, rows: list[dict]) -> str:
        """Genera CSV para actualización manual en Autobits."""
        headers = [
            "proveedor",
            "nit",
            "numero_compra",
            "numero_reserva",
            "numero_documento",
            "valor",
            "fecha",
            "concepto",
            "estado",
        ]
        lines = [",".join(headers)]
        for row in rows:
            values = []
            for key in headers:
                val = row.get(key, "")
                text = "" if val is None else str(val)
                if "," in text or '"' in text:
                    text = '"' + text.replace('"', '""') + '"'
                values.append(text)
            lines.append(",".join(values))
        return "\n".join(lines) + "\n"


def _serialize_cell(value):
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return value


def mapping_to_json(mapping: dict[str, str | None]) -> str:
    return json.dumps({k: v for k, v in mapping.items() if k in AUTOBITS_FIELDS}, ensure_ascii=False)


def mapping_from_json(raw: str | None) -> dict[str, str | None]:
    if not raw:
        return {field: None for field in AUTOBITS_FIELDS}
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return {field: None for field in AUTOBITS_FIELDS}
    return {field: data.get(field) for field in AUTOBITS_FIELDS}
