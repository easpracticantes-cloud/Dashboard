"""Adaptador openpyxl para el Excel de trabajo «CRUCE DE CUENTAS».

A diferencia del export de Autobits, este archivo es una hoja de trabajo
manual: varios proveedores por hoja, cada uno con su bloque de columnas y su
propia fila de encabezados, y el nombre del proveedor escrito arriba del
bloque. El parser localiza esos bloques y devuelve filas normalizadas.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from domain.cruce.fields import (
    CruceParseResult,
    ParsedCruceRow,
    fold,
    header_specificity,
    is_header_text,
    looks_like_invoice,
    looks_like_provider,
    match_block_field,
    to_date,
    to_float,
    to_text,
)

# Un hueco de columna vacía separa bloques laterales (proveedor A | proveedor B).
MAX_BLOCK_GAP = 1
PROVIDER_LOOKBACK = 8
_MESES = {
    "enero",
    "febrero",
    "marzo",
    "abril",
    "mayo",
    "junio",
    "julio",
    "agosto",
    "septiembre",
    "octubre",
    "noviembre",
    "diciembre",
}


class CruceImportError(Exception):
    def __init__(self, message: str, code: str = "CRUCE_IMPORT_ERROR"):
        super().__init__(message)
        self.message = message
        self.code = code


class CruceExcelAdapter:
    def parse(self, path: Path) -> CruceParseResult:
        if not path.exists():
            raise CruceImportError("No se encontró el archivo del cruce de cuentas.", "NOT_FOUND")

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:  # openpyxl lanza varios tipos
            raise CruceImportError(
                "No se pudo leer el Excel de cruce de cuentas. "
                "Guárdelo como .xlsx desde Excel o Google Sheets e intente de nuevo.",
                "UNREADABLE",
            ) from exc

        result = CruceParseResult()
        try:
            for name in wb.sheetnames:
                matrix = [list(row) for row in wb[name].iter_rows(values_only=True)]
                rows = self._parse_sheet(name, matrix)
                result.rows.extend(rows)
                result.sheets.append({"nombre": name, "filas": len(rows)})
        finally:
            wb.close()

        if not result.rows:
            hojas = ", ".join(s["nombre"] for s in result.sheets) or "ninguna"
            raise CruceImportError(
                "No se encontraron filas de cruce en el archivo. "
                f"Hojas revisadas: {hojas}. Debe tener columnas de orden de compra "
                "y FACTURA/CDC o FECHA DE PAGO.",
                "NO_ROWS",
            )

        self._dedupe(result)
        return result

    # -- interno ----------------------------------------------------------

    def _parse_sheet(self, name: str, matrix: list[list]) -> list[ParsedCruceRow]:
        out: list[ParsedCruceRow] = []
        bloques: list[tuple[dict[str, int], str | None]] = []
        tabla: dict[str, int] | None = None
        tabla_proveedor: str | None = None

        for idx, row in enumerate(matrix):
            if self._looks_like_autobits_table(row):
                tabla = self._map_table_columns(row)
                tabla_proveedor = self._table_provider(name, row)
                bloques = []
                continue

            if self._is_header_row(row):
                tabla = None
                bloques = [
                    (block, self._provider_above(matrix, idx, block))
                    for block in self._extract_blocks(row)
                ]
                continue

            if tabla:
                parsed = self._row_from_block(
                    name, idx + 1, row, tabla, tabla_proveedor
                )
                if parsed and not parsed.is_empty():
                    if parsed.nit and fold(parsed.nit) in _MESES:
                        parsed.nit = None
                    out.append(parsed)
                continue

            if not bloques:
                continue

            self._refresh_providers(row, bloques)

            for block, proveedor in bloques:
                parsed = self._row_from_block(name, idx + 1, row, block, proveedor)
                if parsed and not parsed.is_empty():
                    out.append(parsed)
        return out

    def _looks_like_autobits_table(self, row: list) -> bool:
        """VENTAS_DUSTER / CDC BOSQUE: una tabla Autobits, no bloques laterales."""
        folded = [fold(c) for c in row if c is not None]
        joined = " | ".join(folded)
        if "codigo orden de compra" not in joined:
            return False
        return any(
            token in joined
            for token in (
                "nit/cc",
                "nombre proveedor",
                "precio eas",
                "codigo reserva",
            )
        )

    def _map_table_columns(self, row: list) -> dict[str, int]:
        mapping: dict[str, int] = {}
        headers: dict[str, object] = {}
        for idx, cell in enumerate(row):
            if cell is None:
                continue
            campo = match_block_field(cell)
            if not campo:
                continue
            if campo in mapping:
                if header_specificity(campo, cell) > header_specificity(campo, headers[campo]):
                    mapping[campo] = idx
                    headers[campo] = cell
                continue
            mapping[campo] = idx
            headers[campo] = cell
        return mapping

    def _table_provider(self, sheet: str, row: list) -> str | None:
        folded_sheet = fold(sheet)
        if "bosque" in folded_sheet:
            return "PARQUE NATURAL Y CULTURAL BOSQUE DE PALMAS SAS"
        if "duster" in folded_sheet:
            return "VENTAS DUSTER"
        if "luger" in folded_sheet:
            return "PRECOMPRA LUGER"
        return None

    def _is_header_row(self, row: list) -> bool:
        campos = {match_block_field(cell) for cell in row if cell is not None}
        campos.discard(None)
        if "compra" not in campos and "reserva" not in campos:
            return False
        return bool(campos & {"valor", "factura", "pago", "fecha"})

    def _extract_blocks(self, row: list) -> list[dict[str, int]]:
        indexados: list[tuple[int, str, object]] = []
        for idx, cell in enumerate(row):
            if cell is None:
                continue
            campo = match_block_field(cell)
            if campo:
                indexados.append((idx, campo, cell))
        if not indexados:
            return []

        bloques: list[dict[str, int]] = []
        actual: dict[str, int] = {}
        headers: dict[str, object] = {}
        anterior: int | None = None
        for idx, campo, cell in indexados:
            mismo_bloque = anterior is not None and idx - anterior <= MAX_BLOCK_GAP
            if campo in actual and mismo_bloque:
                if header_specificity(campo, cell) > header_specificity(campo, headers.get(campo)):
                    actual[campo] = idx
                    headers[campo] = cell
                anterior = idx
                continue
            corta = anterior is not None and (idx - anterior > MAX_BLOCK_GAP or campo in actual)
            if corta:
                if self._block_is_useful(actual):
                    bloques.append(actual)
                actual = {}
                headers = {}
            actual[campo] = idx
            headers[campo] = cell
            anterior = idx
        if self._block_is_useful(actual):
            bloques.append(actual)
        return bloques

    def _block_is_useful(self, block: dict[str, int]) -> bool:
        if not block:
            return False
        return bool({"compra", "reserva"} & block.keys())

    def _provider_above(
        self,
        matrix: list[list],
        header_idx: int,
        block: dict[str, int],
    ) -> str | None:
        cols = list(block.values())
        desde, hasta = min(cols), max(cols)
        for r in range(header_idx - 1, max(-1, header_idx - PROVIDER_LOOKBACK), -1):
            if r < 0:
                break
            fila = matrix[r]
            for c in range(max(0, desde - 1), min(len(fila), hasta + 2)):
                if looks_like_provider(fila[c]):
                    return to_text(fila[c])
        return None

    def _refresh_providers(
        self,
        row: list,
        bloques: list[tuple[dict[str, int], str | None]],
    ) -> None:
        """Un nombre suelto a media hoja abre una sección de otro proveedor."""
        for c, cell in enumerate(row):
            if cell is None or not looks_like_provider(cell):
                continue
            nombre = to_text(cell)
            for i, (block, _) in enumerate(bloques):
                cols = list(block.values())
                if min(cols) - 1 <= c <= max(cols) + 1:
                    bloques[i] = (block, nombre)

    def _row_from_block(
        self,
        sheet: str,
        row_number: int,
        row: list,
        block: dict[str, int],
        proveedor: str | None,
    ) -> ParsedCruceRow | None:
        def get(campo: str):
            idx = block.get(campo)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        compra = to_text(get("compra"))
        reserva = to_text(get("reserva"))
        valor = to_float(get("valor"))
        factura = to_text(get("factura"))
        pago = get("pago")

        # Encabezado repetido leído como dato
        if is_header_text("compra", compra):
            return None
        if is_header_text("reserva", reserva):
            return None
        if is_header_text("factura", factura):
            factura = None

        if not compra and not reserva:
            return None

        prov_celda = to_text(get("proveedor"))
        if prov_celda and looks_like_provider(prov_celda):
            proveedor = prov_celda
        if proveedor and looks_like_invoice(proveedor):
            proveedor = None

        def celda(campo: str) -> str | None:
            idx = block.get(campo)
            if idx is None:
                return None
            return f"{sheet}!{get_column_letter(idx + 1)}{row_number}"

        return ParsedCruceRow(
            sheet=sheet,
            row_number=row_number,
            proveedor=proveedor,
            nit=to_text(get("nit")),
            numero_compra=compra,
            numero_reserva=reserva,
            concepto=to_text(get("concepto")),
            valor=valor,
            factura_cdc=factura,
            fecha_pago=to_date(pago) or self._texto_pago(pago),
            fecha_ejecucion=to_date(get("fecha")),
            estado_compra=to_text(get("estado")),
            observaciones=to_text(get("observaciones")),
            celda_factura=celda("factura"),
            celda_pago=celda("pago"),
            celda_compra=celda("compra"),
        )

    def _texto_pago(self, value) -> str | None:
        """Algunas celdas de FECHA DE PAGO traen notas ('efectivo', 'pendiente')."""
        text = to_text(value)
        if not text:
            return None
        if is_header_text("pago", text):
            return None
        folded = fold(text)
        if folded in {"-", "--", "x", "n/a", "na", "pendiente", "sin pagar", "por pagar"}:
            return None
        return text if len(text) <= 64 else None

    def _dedupe(self, result: CruceParseResult) -> None:
        vistas: set[tuple] = set()
        unicas: list[ParsedCruceRow] = []
        for row in result.rows:
            clave = (
                fold(row.numero_compra),
                fold(row.numero_reserva),
                row.valor,
                fold(row.factura_cdc),
                row.fecha_pago,
            )
            if clave in vistas:
                result.skipped_rows += 1
                continue
            vistas.add(clave)
            unicas.append(row)
        result.rows = unicas
